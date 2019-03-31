import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
rs = openpyxl.load_workbook('s.xlsx')
wd = openpyxl.load_workbook('d.xlsx')
s_sheet = rs.get_sheet_by_name('Sheet1')
d_sheet = wd.get_sheet_by_name('Sheet1')
dmax_col = d_sheet.max_column
maxrow = s_sheet.max_row
print(s_sheet.max_column)
print((s_sheet['1']))
srcdict = {}
for i in range(2,maxrow+1):
    tmp = []
    city = s_sheet.cell(row=i, column=2).value
    cost = s_sheet.cell(row=i, column=3).value
    tmp.append(city)
    tmp.append(cost)
    srcdict[city]=tmp
for key,value in srcdict.items():
    print(key)
    print(value)
count = 0
for i in range(1,dmax_col,9):
    tagcity = d_sheet.cell(row=1,column=i).value.strip()
    if tagcity == None:
        continue
    elif tagcity == '汇总':
        break
    else:
        d_sheet.cell(row=4,column=i+3,value=srcdict[tagcity][1])
        print(d_sheet.cell(row=4,column=i+3).value)
wd.save('d.xlsx')